import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options




def main():

    url = 'https://odibets.com/odileague'

    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')


    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()),options=chrome_options)

    driver.get(url)

    # Wait for the page to load completely
    wait = WebDriverWait(driver, 10)

    # Locate the "Results" tab element (use the correct selector for your case)
    # Example: Using the text of the tab
    results_tab = wait.until(EC.element_to_be_clickable((By.XPATH, '//li[text()="Results"]')))

    # Click the "Results" tab
    results_tab.click()

    time.sleep(10)
    page_html = driver.page_source

    soup = BeautifulSoup(page_html, 'html.parser')


    driver.quit()

    virtual_cont_soup = soup.find_all("body")[0].find_all("div",class_="page theme-1")[0].find_all("div",class_="virtual")[0].find_all("div",class_="page-container")[0].find_all("div",class_="virtual-container")[0]

    virtual_game_soup = virtual_cont_soup.find_all("div",class_="virtual-main")[0].find_all("div",class_="page-grid")[0].find_all("div",class_="virtual-games")[0]

    results_soup= virtual_game_soup.find_all("div",class_="ba")[0].find_all("div",class_="bc")[0].find_all("div",class_="virtual-rs")[0]

    matches = []

    for match_div in results_soup.find_all('div', class_='rs'):
        
        league = match_div.find_all("div",class_="t")[0].text.split("-")[0]
        match_day =match_div.find_all("div",class_="t")[0].text.split("-")[1]
        for game_div in match_div.find_all("div",class_="rs-g"):
            team1=game_div.find_all("div",class_="g-t")[0].text
            team2=game_div.find_all("div",class_="g-t")[1].text
            score1=game_div.find_all("div",class_="g-s")[0].text[0]
            score2=game_div.find_all("div",class_="g-s")[0].text[1]
        
        
        
            matches.append({
                'League': league,
                'League No': match_day,
                'Home Team': team1,
                'Away Team' :team2,
                'Home Score' : score1,
                'Away Score' : score2
                
                
                
            })

    df = pd.DataFrame(matches)
    df['Game Week'] = df['League'].str[-3:]
    df['Home Score'] = df['Home Score'].astype(int)
    df['Away Score'] = df['Away Score'].astype(int)
    df['Ov 2.5'] = df.apply(lambda row: "YES" if (row['Home Score'] + row['Away Score']) > 2.5 else "NO",axis=1)
    df['GG'] = df.apply(lambda row: "GG" if ((row['Home Score'] > 0 ) and row['Away Score'] > 0) else "NG",axis=1)
    df['GameWeek_ID'] = df['League No'] +df['Game Week']

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'/home/ubuntu/Games_Predictor/Results/Odi_league_{timestamp}.xlsx' 
    df.to_excel(filename, sheet_name="sheet1", index=False)




# def append_to_excel(filename, df, sheet_name='Sheet1'):
#         if os.path.isfile(filename):
#             book = load_workbook(filename)
#             writer = pd.ExcelWriter(filename, engine='openpyxl')
#             writer.book = book
#             if sheet_name in book.sheetnames:
#                 startrow = book[sheet_name].max_row
#             else:
#                 startrow = 0
#             df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, header=(startrow == 0), index=False)
#             writer.save()
#         else:
#             df.to_excel(filename, sheet_name=sheet_name, index=False)



if __name__ == "__main__":
    df = main()
    # append_to_excel(r"C:\Users\User\Documents\IPTV\Odi_league.xlsx",df,sheet_name="Sheet1")

